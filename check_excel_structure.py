#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 파일 구조 확인 스크립트
"""

try:
    import openpyxl
except ImportError:
    print("openpyxl 라이브러리가 설치되지 않았습니다.")
    print("설치 명령어: pip install openpyxl")
    exit(1)

import os

def check_excel_structure():
    """Excel 파일 구조 확인"""
    excel_path = "c:\\Projects\\percenty_project\\percenty_id.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Excel 파일을 찾을 수 없습니다: {excel_path}")
        return
    
    try:
        # Excel 파일 로드
        wb = openpyxl.load_workbook(excel_path)
        print(f"=== Excel 파일 구조 분석: {excel_path} ===")
        print(f"시트 목록: {wb.sheetnames}")
        
        # 활성 시트 분석
        ws = wb.active
        print(f"\n활성 시트: {ws.title}")
        print(f"최대 행: {ws.max_row}")
        print(f"최대 열: {ws.max_column}")
        
        # 헤더 행 확인 (첫 번째 행)
        print("\n=== 헤더 행 (첫 번째 행) ===")
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
            print(f"열 {col} ({chr(64+col)}): {cell_value}")
        
        # 데이터 샘플 확인 (처음 5행)
        print("\n=== 데이터 샘플 (처음 5행) ===")
        for row in range(2, min(7, ws.max_row + 1)):
            print(f"\n행 {row}:")
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                col_letter = chr(64+col)
                print(f"  {col_letter}{row}: {cell_value}")
        
        # 슬래시가 포함된 셀 찾기
        print("\n=== 슬래시(/) 포함 셀 검색 ===")
        slash_cells = []
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and '/' in cell_value:
                    col_letter = chr(64+col)
                    slash_cells.append((f"{col_letter}{row}", cell_value))
        
        if slash_cells:
            print(f"슬래시 포함 셀 {len(slash_cells)}개 발견:")
            for cell_ref, value in slash_cells[:10]:  # 처음 10개만 표시
                print(f"  {cell_ref}: {value}")
            if len(slash_cells) > 10:
                print(f"  ... 및 {len(slash_cells) - 10}개 더")
        else:
            print("슬래시 포함 셀이 없습니다.")
        
        # H, I, J 열 데이터 분석 (이미지 관련 열)
        print("\n=== H, I, J 열 데이터 분석 (이미지 관련) ===")
        image_columns = ['H', 'I', 'J']
        for col_letter in image_columns:
            col_num = ord(col_letter) - 64
            if col_num <= ws.max_column:
                print(f"\n{col_letter}열 데이터:")
                unique_values = set()
                for row in range(2, min(ws.max_row + 1, 20)):  # 처음 18행만 확인
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value:
                        unique_values.add(str(cell_value))
                
                print(f"  고유값들: {sorted(unique_values)}")
        
        wb.close()
        
    except Exception as e:
        print(f"Excel 파일 읽기 중 오류: {e}")

if __name__ == "__main__":
    check_excel_structure()